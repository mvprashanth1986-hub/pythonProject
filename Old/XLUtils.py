import openpyxl

def read_data(file,sheetname,row_count,col_count):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=row_count,column=col_count).value

def write_date(file,sheetname,row_count,col_count,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=row_count,column=col_count).value = data
    workbook.save(file)

def row_count(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return(sheet.max_row)

def col_count(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return(sheet.max_column)